from django.shortcuts import render, redirect, get_object_or_404
from .forms import PaymentForm, SalaryFilterForm
from .models import SalaryInformation, PaymentHistory, Employee
from datetime import datetime
from django.urls import reverse
from django.db import transaction
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.views.decorators.csrf import csrf_protect

def calculate_total_salary(employee):
    return (
        employee.employee_side.hourly_salary * 210 +
        employee.employee_side.overtime_salary * 0 +
        employee.employee_side.the_right_of_the_child * employee.num_children +
        employee.employee_side.ben_kargari +
        employee.employee_side.right_to_housing +
        employee.employee_side.base_years

    ) 

def calculate_salary(request):
    current_user = request.user
    employees = Employee.objects.filter(user=current_user)
    calculated_salaries = []
    missing_employee_side_list = []
    form = SalaryFilterForm(request.GET)
    if form.is_valid():
        employee_id = form.cleaned_data.get('employee_id')
        month = form.cleaned_data.get('month')
        month2 = form.cleaned_data.get('month2')
        
        if employee_id:
            employees = employees.filter(id=employee_id.id)

        try:
            start_date = datetime.strptime(month, '%Y-%m') if month else None
            end_date = datetime.strptime(month2, '%Y-%m') if month2 else None
        except ValueError:
            start_date = end_date = None

        if not start_date:
            start_date = datetime.now().replace(day=1)
        if not end_date:
            end_date = start_date

    else:
        start_date = datetime.now().replace(day=1)
        end_date = start_date

    with transaction.atomic():
        for employee in employees:
            if start_date and end_date:
                current_date = start_date
                while current_date <= end_date:
                    salary_info = SalaryInformation.objects.filter(
                        employee=employee,
                        salary_month=current_date.strftime('%Y-%m-%d')
                    ).first()
                    if salary_info and employee.employee_side is not None:
                        calculated_salaries.append(salary_info)
                    current_date = add_months(current_date, 1)
            else:
                if employee.employee_side is None:
                    missing_employee_side_list.append({'id': employee.id, 'name': employee.name})
                else:
                    total_salary = calculate_total_salary(employee)
                    salary_info, created = SalaryInformation.objects.get_or_create(
                        employee=employee,
                        salary_month=start_date,
                        defaults={
                            'monthly_income': total_salary,
                            'monthly_expenses': 0,
                        }
                    )
                    calculated_salaries.append(salary_info)

    context = {
        'calculated_salaries': calculated_salaries,
        'missing_employee_side_list': missing_employee_side_list,
        'current_month': start_date.strftime('%Y-%m'),
        'month2': month2,
        'month_filter': month,
        'employee_filter': employee_id,
        'form': form,
    }
    return render(request, 'salary_list.html', context)

def add_months(source_date, months):
    month = source_date.month - 1 + months
    year = source_date.year + month // 12
    month = month % 12 + 1
    day = min(source_date.day, 28)  # Handling the edge case for February
    return datetime(year, month, day)

def add_months_excel(sourcedate, months):
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, [31,
        29 if year % 4 == 0 and not year % 100 == 0 or year % 400 == 0 else 28,
        31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return datetime(year, month, day)

@csrf_protect
def export_to_excel(request):
    if request.method == "POST":
        current_user = request.user
        employees = Employee.objects.filter(user=current_user)
        missing_employee_side_list = []

        form = SalaryFilterForm(request.POST)
        if form.is_valid():
            employee_id = form.cleaned_data.get('employee_id')
            month = form.cleaned_data.get('month')
            month2 = form.cleaned_data.get('month2')
            
            if employee_id:
                employees = employees.filter(id=employee_id.id)

            try:
                start_date = datetime.strptime(month, '%Y-%m') if month else None
                end_date = datetime.strptime(month2, '%Y-%m') if month2 else None
            except ValueError:
                start_date = end_date = None

            if not start_date:
                start_date = datetime.now().replace(day=1)
            if not end_date:
                end_date = start_date

        else:
            start_date = datetime.now().replace(day=1)
            end_date = start_date

        with transaction.atomic():
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Salaries'

            # Write header row
            headers = ['Employee ID', 'Employee Name', 'Employee Side', 'Salary Month', 'Monthly Income', 'Monthly Expenses', 'Payment Details', 'Payment Date Details']
            worksheet.append(headers)

            # Apply header styles
            header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
            header_font = Font(bold=True, color="403151")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(left=Side(border_style="thin", color="8DB4E2"),
                                    right=Side(border_style="thin", color="8DB4E2"),
                                    top=Side(border_style="thin", color="8DB4E2"),
                                    bottom=Side(border_style="thin", color="8DB4E2"))
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border

            alt_fill1 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            alt_fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            row_counter = 2
            color_counter = 0

            for employee in employees:
                if start_date and end_date:
                    current_date = start_date
                    while current_date <= end_date:
                        salary_info = SalaryInformation.objects.filter(
                            employee=employee,
                            salary_month=current_date.strftime('%Y-%m-%d')
                        ).first()
                        if salary_info and employee.employee_side is not None:
                            # Calculate total monthly expenses from PaymentHistory
                            payment_history_entries = PaymentHistory.objects.filter(salary_information=salary_info)
                            total_expenses = sum(entry.amount for entry in payment_history_entries)

                            # Prepare data rows
                            payment_details = [entry.amount for entry in payment_history_entries]
                            payment_dates = [entry.payment_date for entry in payment_history_entries]
                            max_len = max(len(payment_details), len(payment_dates), 1)

                            for i in range(max_len):
                                row = [
                                    salary_info.employee.id if i == 0 else "",
                                    salary_info.employee.name if i == 0 else "",
                                    str(salary_info.employee.employee_side) if i == 0 else "",  # Ensure employee_side is converted to string
                                    salary_info.salary_month if i == 0 else "",
                                    salary_info.monthly_income if i == 0 else "",
                                    salary_info.monthly_expenses if i == 0 else "",
                                    payment_details[i] if i < len(payment_details) else "None",
                                    payment_dates[i] if i < len(payment_dates) else "None"
                                ]
                                worksheet.append(row)

                                # Apply alternating fill to all columns
                                fill = alt_fill1 if color_counter % 2 == 0 else alt_fill2
                                for col in range(1, worksheet.max_column + 1):
                                    worksheet.cell(row=row_counter, column=col).fill = fill
                                row_counter += 1

                            # Merge cells
                            if max_len > 1:  # Check if there are multiple rows to merge
                                for col in "ABCDEF":
                                    worksheet.merge_cells(f"{col}{worksheet.max_row - max_len + 1}:{col}{worksheet.max_row}")

                            color_counter += 1

                        current_date = add_months_excel(current_date, 1)

            # Apply borders to all cells
            thin_border = Border(left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000"))
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border

            # Prepare response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=salaries.xlsx'
            
            workbook.save(response)
            return response

    return HttpResponse(status=405)  # Method not allowed

def salary_pay(request, SalaryInformation_id):
    """
    View for creating a payment for a specific SalaryInformation instance.

    If the request is a POST request and the form is valid, create a new
    PaymentHistory instance with the data from the form and redirect to the
    calculate_salary view.

    If the request is a GET request or the form is not valid, render the
    payment_form.html template with the PaymentForm instance.

    Args:
        request (django.http.HttpRequest): The request sent to the server.
        SalaryInformation_id (int): The ID of the SalaryInformation instance
            to make a payment for.

    Returns:
        django.http.HttpResponse: The response returned to the user.
    """
    salary_info = get_object_or_404(SalaryInformation, pk=SalaryInformation_id)
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            amount = form.cleaned_data['amount']
            payment_date = form.cleaned_data['payment_date']
            payment = PaymentHistory.objects.create(
                salary_information=salary_info,
                amount=amount,
                payment_date=payment_date)
            salary_info.monthly_expenses += amount
            salary_info.save()
            return redirect('calculate_salary')
    else:
        form = PaymentForm()
    context = {'form': form}
    return render(request, 'payment_form.html', context)

def monthly_payment_history(request, employee_id, month):
    """
    View for displaying the payment history for a specific employee in a
    particular month.

    Args:
        request (django.http.HttpRequest): The request sent to the server.
        employee_id (int): The ID of the employee to display payment history
            for.
        month (str): The month to display payment history for, in the format
            '%B'. For example, 'January', 'February', etc.

    Returns:
        django.http.HttpResponse: The response returned to the user.
    """
    employee = get_object_or_404(Employee, id=employee_id)
    month_number = datetime.strptime(month, '%B').month
    salary_info_for_month = SalaryInformation.objects.filter(
        employee_id=employee_id,
        salary_month__month=month_number)
    payment_history = PaymentHistory.objects.filter(
        salary_information__in=salary_info_for_month)
    context = {
        'employee': employee,
        'payment_history': payment_history,
        'month': month,
    }
    return render(request, 'monthly_payment_history.html', context)

def delete_payment(request, payment_id):
    """
    View for deleting a payment from the database

    Args:
        request (django.http.HttpRequest): The request sent to the server.
        payment_id (int): The ID of the payment to delete.

    Returns:
        django.http.HttpResponse: The response returned to the user.
    """
    payment = get_object_or_404(PaymentHistory, pk=payment_id)
    # Get the amount of the payment that is being deleted
    month_expense = payment.amount 
    # Delete the payment from the database
    payment.delete()
    # Get the salary information object that the payment is associated with
    salary_info = payment.salary_information
    # Subtract the amount of the deleted payment from the monthly expenses
    salary_info.monthly_expenses -= month_expense
    # Save the updated salary information
    salary_info.save()
    # Get the URL for the monthly payment history view
    month_url = reverse('monthly_payment_history',
        kwargs={'employee_id': salary_info.employee_id,
            'month': salary_info.salary_month.strftime('%B').lower()})
    # Redirect the user to the monthly payment history view
    return redirect(month_url)